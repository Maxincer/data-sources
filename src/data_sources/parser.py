#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Parse raw exchange data files into unified daily bar records.

Each parser function reads a raw file from data/raw/structured/ and yields
dicts with keys matching t_futures_info columns.

Stat tracking: total_records, missing count per field, missing deviation.
"""

import asyncio
import csv
import json
import os
import re
from pathlib import Path
import aiohttp
from bs4 import BeautifulSoup

from data_sources.modifier import (
    safe_ceil,
    safe_floor,
    calc_shfe_ine_limit_prices,
    zero_price_to_none,
    _is_shfe_product,
    _is_ine_product,
)

RAW_DIR = Path(os.environ["DATA_DIR"]) / "raw" / "structured"

# Exchange code suffixes for DB
SUFFIX_MAP = {
    "CFE": ".CFE",
    "CSI": ".CSI",
    "CZC": ".CZC",
    "DCE": ".DCE",
    "GFE": ".GFE",
    "INE": ".INE",
    "SHF": ".SHF",
}

# DB columns in order
DB_COLS = [
    "code", "date", "open", "high", "low", "close",
    "volume", "amt", "oi", "settle", "maxup", "maxdown",
    "if_basis", "long_margin", "short_margin", "minoq", "maxoq",
]


class ParseStats:
    """Track parsing statistics per exchange."""

    def __init__(self, exchange: str, data_type: str):
        self.exchange = exchange
        self.data_type = data_type
        self.total = 0                     # total parsed records
        self.missing: dict[str, int] = {}  # field -> count of missing values
        self.errors: list[str] = []        # error messages

    def add_record(self, record: dict):
        """Count a record and its missing fields."""
        self.total += 1
        for col in DB_COLS:
            val = record.get(col)
            if val is None or (isinstance(val, float) and val != val):
                self.missing[col] = self.missing.get(col, 0) + 1

    def add_error(self, msg: str):
        self.errors.append(msg)

    def merge(self, other: "ParseStats"):
        self.total += other.total
        for col, cnt in other.missing.items():
            self.missing[col] = self.missing.get(col, 0) + cnt
        self.errors.extend(other.errors)

    @property
    def stats_summary(self) -> dict:
        # Fields that are naturally absent for this exchange (not an anomaly)
        _EXPECTED_NULLS = {
            "CFFEX": set(),
            "CZCE": {"if_basis"},
            "DCE": {"if_basis"},
            "SHFE": {"if_basis"},
            "INE": {"if_basis"},
            "GFEX": {"if_basis"},
            "CSI": {"if_basis", "long_margin", "short_margin",
         "minoq", "maxoq"},
        }
        expected = _EXPECTED_NULLS.get(self.exchange, set())
        return {
            "exchange": self.exchange,
            "data_type": self.data_type,
            "total_records": self.total,
            "missing": {k: v for k, v in self.missing.items()
                        if v > 0 and k not in expected},
            "errors": len(self.errors),
        }


# -------------------------------------------------------------------------
# Verification helpers
# -------------------------------------------------------------------------

def verify_price_order(record: dict) -> str | None:
    """Check open/high/low/close price ordering."""
    o, h, l, c = [record.get(k) for k in ("open", "high", "low", "close")]
    if o is not None and h is not None and l is not None:
        if l > h:
            return "low exceeds high"
        if o is not None and (o > h or o < l):
            return "open outside high/low range"
        if c is not None and (c > h or c < l):
            return "close outside high/low range"
    return None


def verify_settle_vs_close(record: dict) -> str | None:
    """Check settle is close to close (within 5%)."""
    s, c = record.get("settle"), record.get("close")
    if s is not None and c is not None and c != 0:
        dev = abs(s - c) / abs(c) * 100
        if dev > 5:
            return f"settle deviates {dev:.2f}% from close"
    return None


# -------------------------------------------------------------------------
# Exchange-specific parsers
# -------------------------------------------------------------------------

def parse_directory(dirpath: Path) -> list[dict]:
    """Auto-discover and parse all raw files in data/raw/structured/."""
    records = []
    for fpath in sorted(dirpath.iterdir()):
        if fpath.suffix == ".jsonl":
            continue
        try:
            parsed = parse_file(fpath)
            records.extend(parsed)
        except Exception:
            continue
    return records


def parse_file(fpath: Path) -> list[dict]:
    """
    Parse a single raw file and return list of record dicts.
    If verifier is provided, run business-logic checks and log issues.
    """
    name = fpath.name
    if "CFFEX" in name and "MarketData" in name:
        return _parse_cffex_market(fpath)
    if "CFFEX" in name and "Settlement" in name:
        return _parse_cffex_settlement(fpath)
    if ("SHFE" in name or "INE" in name) and "DailyMarketData" in name:
        return _parse_shfe_ine_daily(fpath)
    if ("SHFE" in name or "INE" in name) and "Settlement" in name:
        return _parse_shfe_ine_settlement(fpath)
    if "CZCE" in name and "DailyMarketData" in name:
        return _parse_czce_daily(fpath)
    if "CZCE" in name and "Settlement" in name:
        return _parse_czce_settlement(fpath)
    if "CZCE" in name and "TradingParameters" in name:
        return _parse_czce_tradepara(fpath)
    if "DCE" in name and "DailyMarketData" in name:
        return _parse_dce_daily(fpath)
    if "DCE" in name and "Settlement" in name:
        return _parse_dce_settlement(fpath)
    if "GFEX" in name and "DailyMarketData" in name:
        return _parse_gfex_daily(fpath)
    if "GFEX" in name and "Settlement" in name:
        return _parse_gfex_settlement(fpath)
    # Trading parameters
    if "CFFEX" in name and "TradingParameters" in name:
        return _parse_cffex_tradepara(fpath)
    if "SHFE" in name and "TradingParameters" in name:
        return _parse_shfe_ine_tradepara(fpath, ".SHF")
    if "INE" in name and "TradingParameters" in name:
        return _parse_shfe_ine_tradepara(fpath, ".INE")
    if "DCE" in name and "TradingParameters" in name and "Variety" not in name:
        return _parse_dce_tradepara(fpath)
    if "DCE" in name and "VarietyTradingParam" in name:
        return _parse_dce_tradingparam(fpath)
    if "GFEX" in name and "TradingParameters" in name:
        return _parse_gfex_tradepara(fpath)
    if "CSI" in name and "MarketData" in name:
        return _parse_csi_market(fpath)
    return []


def _exchange_suffix(exchange: str) -> str:
    """Map exchange name to DB code suffix."""
    m = {"SHFE": "SHF", "CFFEX": "CFE", "CZCE": "CZC",
         "DCE": "DCE", "GFEX": "GFE", "INE": "INE"}
    return SUFFIX_MAP.get(m.get(exchange, exchange), f".{exchange}")


# -----------------------------------------------------------------------
# CFFEX market CSV (GBK, from hqsj)
# -----------------------------------------------------------------------

def _parse_cffex_market(fpath: Path) -> list[dict]:
    """Parse CFFEX daily market data CSV (GBK)."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="gbk", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get("合约代码", "").strip()
            if not code or code == "小计" or "总计" in code or "合计" in code:
                continue
            if any(kw in code for kw in ["-C-", "-P-"]):
                continue  # skip options
            # Keep only IF/IC/IH index futures
            if not code.startswith(("IF","IC","IH","IM","T","TF","TL","TS")):
                continue
            rec = {
                "code": code + ".CFE",
                "date": date,
                "open": _float(row.get("今开盘")),
                "high": _float(row.get("最高价")),
                "low": _float(row.get("最低价")),
                "close": _float(row.get("今收盘")),
                "volume": _float(row.get("成交量")),
                "amt": _float(row.get("成交金额")),
                "oi": _float(row.get("持仓量")),
                "settle": _float(row.get("今结算")),
            }
            if rec["amt"] is not None:
                rec["amt"] = rec["amt"] * 1e4  # 亿元->元
            records.append(rec)
    return records


# -----------------------------------------------------------------------
# CFFEX settlement CSV (GBK, from jscs)
# -----------------------------------------------------------------------

def _parse_cffex_settlement(fpath: Path) -> list[dict]:
    """Parse CFFEX settlement parameters CSV (GBK).

    Line 0 is a title/date row, line 1 is actual headers.
    Columns: 期货合约,合约多头保证金标准,合约空头保证金标准,...
    """
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="gbk", errors="replace") as f:
        lines = f.read().strip().split("\n")
    reader = csv.DictReader(lines[1:])
    for row in reader:
        code = row.get("期货合约", "").strip()
        if not code:
            continue
        if not code.startswith(("IF","IC","IH","IM","T","TF","TL","TS")):
            continue
        lm = row.get("合约多头保证金标准", "")
        sm = row.get("合约空头保证金标准", "")
        records.append({
            "code": code + ".CFE",
            "date": date,
            "long_margin": _pct_val(lm),
            "short_margin": _pct_val(sm),
        })
    return records


# -----------------------------------------------------------------------
# SHFE / INE daily market data JSON
# -----------------------------------------------------------------------

def _parse_shfe_ine_daily(fpath: Path) -> list[dict]:
    """Parse SHFE/INE kx*.dat daily market data JSON."""
    records = []
    date = fpath.name.split(".")[0]
    name = fpath.name
    exchange_code = "SHF" if "SHFE" in name else "INE"
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("o_curinstrument", []):
        pgid = item.get("PRODUCTGROUPID", "")
        dmonth = str(item.get("DELIVERYMONTH", ""))
        if not pgid or not dmonth:
            continue
        pgid_up = pgid.upper()
        if "小计" in dmonth or "TAS" in dmonth.upper():
            continue
        if pgid_up.endswith("_TAS"):
            pgid_up = pgid_up.replace("_TAS", "")
            dmonth = dmonth + "TAS"
        if _is_ine_product(pgid_up) and exchange_code == "SHF":
            continue
        if _is_shfe_product(pgid_up) and exchange_code == "INE":
            continue
        code = f"{pgid_up}{dmonth}.{exchange_code}"
        rec = {
            "code": code,
            "date": date,
            "open": _float(item.get("OPENPRICE")),
            "high": _float(item.get("HIGHESTPRICE")),
            "low": _float(item.get("LOWESTPRICE")),
            "close": _float(item.get("CLOSEPRICE")),
            "volume": _float(item.get("VOLUME")),
            "amt": _float(item.get("TURNOVER")),
            "oi": _float(item.get("OPENINTEREST")),
            "settle": _float(item.get("SETTLEMENTPRICE")),
            "pre_settle": _float(item.get("PRESETTLEMENTPRICE")),
        }
        if rec["amt"] is not None:
            rec["amt"] = rec["amt"] * 1e4
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# SHFE / INE settlement JSON# SHFE / INE settlement JSON
# -----------------------------------------------------------------------

def _parse_shfe_ine_settlement(fpath: Path) -> list[dict]:
    """Parse SHFE/INE js*.dat settlement JSON."""
    records = []
    date = fpath.name.split(".")[0]
    name = fpath.name
    exchange_code = "SHF" if "SHFE" in name else "INE"
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("o_cursor", []):
        instr = item.get("INSTRUMENTID", "")
        if not instr or "小计" in instr:
            continue
        instr_up = instr.upper()
        if (
            exchange_code == "SHF"
            and instr_up[:2] in {"SC","BC","LU","NR","EC"}
        ):
            continue
        if (
            exchange_code == "INE"
            and instr_up[:2] in {
                "CU","AL","ZN","PB","NI","SN",
                "AU","AG","RB","WR","HC","SS","FU","BU",
                "BR","RU","SP","OP","AD","AO"
            }
        ):
            continue
        code = instr_up + f".{exchange_code}"
        records.append({
            "code": code,
            "date": date,
            "settle": _float(item.get("SETTLEMENTPRICE")),
            "long_margin": _float(item.get("SPECLONGMARGINRATIO")),
            "short_margin": _float(item.get("SPECSHORTMARGINRATIO")),
        })
    return records


# -----------------------------------------------------------------------
# CZCE daily market data TXT
# -----------------------------------------------------------------------

def _parse_czce_daily(fpath: Path) -> list[dict]:
    """Parse CZCE daily market data pipe-delimited TXT."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        lines = f.read().strip().split("\n")
        for line in lines[2:]:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) < 14:
                continue
            code_raw = parts[0]
            if not code_raw or not re.match(r"^[A-Z]+\d+", code_raw):
                continue
            rec = {
                "code": code_raw + ".CZC",
                "date": date,
                "pre_settle": _float(parts[1]),
                "open": _float(parts[2]),
                "high": _float(parts[3]),
                "low": _float(parts[4]),
                "close": _float(parts[5]),
                "settle": _float(parts[6]),
                "volume": _float(parts[9]),
                "oi": _float(parts[10]),
                "amt": _float(parts[12]),
            }
            if rec["amt"] is not None:
                rec["amt"] = rec["amt"] * 1e4
            records.append(rec)
            records.append(rec)
    return records


def _parse_czce_settlement(fpath: Path) -> list[dict]:
    """Parse CZCE settlement parameters TXT (UTF-8).

    Columns: 合约代码|当日结算价|是否单边市|连续单边市天数|交易保证金率(%)|涨跌停板(%)
    """
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        lines = f.read().strip().split("\n")
        for line in lines[2:]:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) < 6:
                continue
            code_raw = parts[0]
            if not code_raw or not re.match(r"^[A-Z]+\d+", code_raw):
                continue
            settle = _float(parts[1])
            margin_pct = (
                    float(parts[4].replace(",","")) / 100.0
                    if parts[4].strip().replace(",","") else None
                )
            rec = {
                "code": code_raw + ".CZC",
                "date": date,
                "settle": settle,
                "long_margin": margin_pct,
                "short_margin": margin_pct,
            }
            records.append(rec)
    return records


# -----------------------------------------------------------------------
# CZCE trading parameters TXT (FutureTradeParam.txt)
# Columns: 合约代码|上市日期|交易单位|最小变动价位|上日结算价|
#          涨跌停板幅度(%)|日持仓限额|交易限额|最小开仓量下单量|
#          限价单最大下单量|市价单最大下单量
# -----------------------------------------------------------------------

def _parse_czce_tradepara(fpath: Path) -> list[dict]:
    """Parse CZCE FutureTradeParam.txt (UTF-8)."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        lines = f.read().strip().split("\n")
    for line in lines[2:]:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 11:
            continue
        code_raw = parts[0]
        if not code_raw or not re.match(r"^[A-Z]+\d+", code_raw):
            continue
        limit_raw = parts[5]
        tick = _float(parts[3])
        pre_settle = _float(parts[4])
        limit_pct = _float(limit_raw) / 100.0 if limit_raw else None
        minoq = _float_to_int(_float(parts[8]))
        maxoq = _float_to_int(_float(parts[9]))  # 限价单最大下单量
        rec = {
            "code": code_raw + ".CZC",
            "date": date,
            "pre_settle": pre_settle,
            "_limit_pct": limit_pct,
            "_tick_size": tick,
            "minoq": minoq,
            "maxoq": maxoq,
        }
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# DCE daily market data JSON
# -----------------------------------------------------------------------

def _parse_dce_daily(fpath: Path) -> list[dict]:
    """Parse DCE daily market data JSON."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("data", []):
        cid = item.get("contractId", "")
        if not cid:
            continue
        rec = {
            "code": cid + ".DCE",
            "date": date,
            "open": _float(item.get("open")),
            "high": _float(item.get("high")),
            "low": _float(item.get("low")),
            "close": _float(item.get("close")),
            "volume": _float(item.get("volumn")),
            "amt": _float(item.get("turnover")),
            "oi": _float(item.get("openInterest")),
            "settle": _float(item.get("clearPrice")),
            "_last_clear": _float(item.get("lastClear")),
        }
        if rec["amt"] is not None:
            rec["amt"] = rec["amt"] * 1e4  # 万元->元
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# DCE settlement JSON
# -----------------------------------------------------------------------

def _parse_dce_settlement(fpath: Path) -> list[dict]:
    """Parse DCE settlement parameters JSON."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("data", []):
        cid = item.get("contractId", "")
        if not cid:
            continue
        records.append({
            "code": cid + ".DCE",
            "date": date,
            "settle": _float(item.get("clearPrice")),
            "long_margin": _float(item.get("specBuyRate")),
            "short_margin": _float(item.get("specSellRate")),
        })
    return records


# -----------------------------------------------------------------------
# GFEX daily market data JSON
# -----------------------------------------------------------------------

def _parse_gfex_daily(fpath: Path) -> list[dict]:
    """Parse GFEX daily market data JSON."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("data", []):
        variety = item.get("variety", "")
        if not variety or variety == "总计":
            continue
        deliv = str(item.get("delivMonth", "")).strip()
        if not deliv:
            continue
        var_order = item.get("varietyOrder", "").strip()
        code = f"{var_order}{deliv}.GFE" if var_order else None
        if not code:
            continue
        rec = {
            "code": code,
            "date": date,
            "open": _float(item.get("open")),
            "high": _float(item.get("high")),
            "low": _float(item.get("low")),
            "close": _float(item.get("close")),
            "volume": _float(item.get("volumn")),
            "amt": _float(item.get("turnover")),
            "oi": _float(item.get("openInterest")),
            "settle": _float(item.get("clearPrice")),
        }
        if rec["amt"] is not None:
            rec["amt"] = rec["amt"] * 1e4  # 万元->元
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# GFEX settlement JSON
# -----------------------------------------------------------------------

def _parse_gfex_settlement(fpath: Path) -> list[dict]:
    """Parse GFEX settlement parameters JSON."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("data", []):
        cid = item.get("contractId", "")
        if not cid:
            continue
        records.append({
            "code": cid + ".GFE",
            "date": date,
            "settle": _float(item.get("clearPrice")),
            "long_margin": _float(item.get("specBuyRate")),
            "short_margin": _float(item.get("specSellRate")),
        })
    return records


# -----------------------------------------------------------------------
# CFFEX trading parameters CSV (jycs)
# Columns: 合约代码,合约月份,挂盘基准价,上市日,最后交易日,
#          涨停板幅度(%),跌停板幅度(%),涨停板价位,跌停板价位,持仓限额
# File is GBK encoded
# -----------------------------------------------------------------------

def _parse_cffex_tradepara(fpath: Path) -> list[dict]:
    """Parse CFFEX trading parameters CSV (GBK)."""
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="gbk", errors="replace") as f:
        lines = f.read().strip().split("\n")

    reader = csv.DictReader(lines[1:])  # skip title line
    for row in reader:
        code = row.get("合约代码", "").strip()
        if not code or "-" in code:
            continue  # skip option contracts
        if not code.startswith(("IF","IC","IH","IM","T","TF","TL","TS")):
            continue
        rec = {
            "code": code + ".CFE",
            "date": date,
            "maxup": _float(row.get("涨停板价位")),
            "maxdown": _float(row.get("跌停板价位")),
        }
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# SHFE/INE trading parameters JSON (ContractDailyTradeArgument)
# Keys: INSTRUMENTID, UPPER_VALUE, LOWER_VALUE, PRICE_LIMITS,
#       SPEC_LONGMARGINRATIO, SPEC_SHORTMARGINRATIO, TRADINGDAY
# -----------------------------------------------------------------------

def _parse_shfe_ine_tradepara(fpath: Path, suffix: str) -> list[dict]:
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("ContractDailyTradeArgument", []):
        instr = item.get("INSTRUMENTID", "")
        if not instr or "小计" in instr:
            continue
        instr_up = instr.upper()
        # 提取品种代码（去掉月份数字）
        pgid = "".join(c for c in instr_up if c.isalpha())
        if suffix == ".SHF" and _is_ine_product(pgid):
            continue
        if suffix == ".INE" and _is_shfe_product(pgid):
            continue
        upper = _float(item.get("UPPER_VALUE"))
        upper = _float(item.get("UPPER_VALUE"))
        lower = _float(item.get("LOWER_VALUE"))
        lm = _float(item.get("SPEC_LONGMARGINRATIO"))
        sm = _float(item.get("SPEC_SHORTMARGINRATIO"))
        rec = {
            "code": instr_up + suffix,
            "date": date,
            "long_margin": lm,
            "short_margin": sm,
            "_limit_pct": upper,
            "_limit_pct_down": lower,
        }
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# DCE trading parameters JSON (dayTradPara API)
# Keys: contractId, riseLimitRate, riseLimit, fallLimit,
#       specBuyRate, selfTotBuyPosiQuota
# -----------------------------------------------------------------------

def _parse_dce_tradepara(fpath: Path) -> list[dict]:
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("data", [])
    for item in items:
        cid = item.get("contractId", "")
        if not cid or "-" in cid or "小计" in cid or "TAS" in cid.upper():
            continue
        rec = {
            "code": cid + ".DCE",
            "date": date,
            "maxup": _float(item.get("riseLimit")),
            "maxdown": _float(item.get("fallLimit")),
            "long_margin": _float(item.get("specBuyRate")),
            "short_margin": _float(
                            item.get("specSellRate")
                            or item.get("specBuyRate")
                        ),
            "_rise_limit_rate": _float(item.get("riseLimitRate")),
            "_pre_settle": _float(item.get("hedgeBuy") or item.get("specBuy")),
        }
        records.append(rec)
    return records


def _parse_dce_tradingparam(fpath: Path) -> list[dict]:
    """Parse DCE product-level tradingParam (varietyId + maxHand).

    Extracts maxHand → maxoq for all DCE products.
    """
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("data", [])
    for item in items:
        vid = item.get("varietyId", "")
        if not vid:
            continue
        max_hand = _float(item.get("maxHand"))
        if max_hand is None or max_hand <= 0:
            continue
        records.append({
            "code": vid + ".DCE",
            "date": date,
            "maxoq": _float_to_int(max_hand),
        })
    return records


# -----------------------------------------------------------------------
# GFEX trading parameters JSON (loadDayList API)
# Keys: contractId, riseLimitRate, riseLimit, fallLimit,
#       specBuyRate, selfTotBuyPosiQuota
# -----------------------------------------------------------------------

def _parse_gfex_tradepara(fpath: Path) -> list[dict]:
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("data", [])
    for item in items:
        cid = item.get("contractId", "")
        if not cid or "-" in cid or "TAS" in cid.upper():
            continue
        rec = {
            "code": cid + ".GFE",
            "date": date,
            "long_margin": _float(item.get("specBuyRate")),
            "short_margin": _float(item.get("specSellRate")),
            "maxup": _float(item.get("riseLimit")),
            "maxdown": _float(item.get("fallLimit")),
        }
        records.append(rec)
    return records


# -----------------------------------------------------------------------
# CSI index market data (JSON API from csindex.com.cn)
# Keys: indexCode, indexName, latestClose
# -----------------------------------------------------------------------

def _parse_csi_market(fpath: Path) -> list[dict]:
    records = []
    date = fpath.name.split(".")[0]
    with open(fpath, encoding="utf-8") as f:
        data = json.load(f)
    for item in data.get("data", []):
        code = item.get("indexCode", "")
        close = _float(item.get("latestClose"))
        if not code:
            continue
        records.append({
            "code": f"{code}.CSI",
            "date": date,
            "close": close,
        })
    return records


# -----------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------

# -----------------------------------------------------------------------
# Tick size lookup (CZCE products)
# 最小变动价位, used for price limit rounding
# -----------------------------------------------------------------------

def _parse_limit_pct(raw: str) -> float | None:
    """Parse CZCE limit pct: '±13' -> 0.13, '±20' -> 0.20."""
    if raw is None:
        return None
    raw = raw.strip().lstrip("±").strip()
    if not raw:
        return None
    try:
        return float(raw.replace(",", "")) / 100.0
    except (ValueError, TypeError):
        return None


def _float_to_int(val: float | None) -> int | None:
    """Convert float to int, or None."""
    return int(val) if val is not None else None


def _float(val) -> float | None:
    """Convert string/None to float or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).replace(",", "").replace(" ", "")
    if not val or val == "null" or val == "None":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _pct_val(val: str) -> float | None:
    """Parse percentage string like '12%' -> 0.12.
    Also handles '12%' (strips %% before float conversion), and plain '12'.
    """
    if val is None:
        return None
    raw = str(val).strip()
    has_pct = "%" in raw
    cleaned = raw.replace("%", "").replace(",", "").replace(" ", "")
    if not cleaned or cleaned in ("null", "None", ""):
        return None
    try:
        v = float(cleaned)
    except (ValueError, TypeError):
        return None
    return v / 100.0 if has_pct else v


# -----------------------------------------------------------------------
# 最小变动价位(从原始 HTML 解析)
# -----------------------------------------------------------------------

PRODUCT_CONFIG_DIR = (
    Path(os.environ["DATA_DIR"]) / "raw" / "product_configs"
)


def _parse_tick_from_html(html: str) -> float | None:
    """从 SHFE/INE 合约规格 HTML 页面提取最小变动价位。

    SHFE/INE 产品页面的合约规格表为 <table class='table_info'>,
    其中"最小变动价位"在 <td> 中,其下一个 <td> 为对应数值。
    """
    soup = BeautifulSoup(html, "html.parser")
    # 定位合约规格表
    table = soup.find("table", class_="table_info")
    if not table:
        return None
    for td in table.find_all(["td", "th"]):
        if td.get_text(strip=True) == "最小变动价位":
            next_td = td.find_next_sibling(["td", "th"])
            if next_td:
                tick_str = next_td.get_text(strip=True)
                num_match = re.search(r"([\d.]+)", tick_str)
                if num_match:
                    return float(num_match.group(1))
    return None


def _load_exchange_map() -> dict[str, str]:
    """从 product_configs 文件名构建交易所归属映射。

    只扫描文件名(YYYYMMDD.EXCHANGE.product_code_f.html)，不打开文件。
    Returns: {pgid: "SHF"|"INE"}
    """
    exchange_map: dict[str, str] = {}
    for fpath in PRODUCT_CONFIG_DIR.iterdir():
        if fpath.suffix != ".html":
            continue
        parts = fpath.stem.split(".")
        if len(parts) < 3:
            continue
        exchange = parts[1]
        if exchange not in ("SHFE", "INE"):
            continue
        code_key = parts[-1].lower()
        if not code_key.endswith("_f"):
            continue
        pgid = code_key.rsplit("_", 1)[0]
        # INE 优先: SHFE 页面可能列出 INE 品种
        if pgid not in exchange_map or exchange == "INE":
            exchange_map[pgid] = "SHF" if exchange == "SHFE" else "INE"
    return exchange_map


def _load_tick_map(date_str: str) -> dict[str, float]:
    tick_map: dict[str, float] = {}
    prefix = f"{date_str}."
    for fpath in PRODUCT_CONFIG_DIR.iterdir():
        if not fpath.name.startswith(prefix) or fpath.suffix != ".html":
            continue
        code_key = fpath.stem.rsplit(".", 1)[-1].lower()
        if not code_key.endswith("_f"):
            continue
        pgid = code_key.rsplit("_", 1)[0]
        tick = _parse_tick_from_html(fpath.read_text(encoding="utf-8"))
        if tick is not None:
            tick_map[pgid] = tick
    return tick_map


# -----------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------


# -----------------------------------------------------------------------
# Orchestration
# -----------------------------------------------------------------------

def parse_all():
    """Parse all raw files and return (records, stats_summary)."""
    all_records: list[dict] = []
    all_stats = []
    for fpath in sorted(RAW_DIR.iterdir()):
        if not fpath.is_file() or fpath.suffix == ".jsonl":
            continue
        name = fpath.name
        exchange = name.split(".")[1] if len(name.split(".")) > 1 else "?"
        data_type = name.split(".")[2] if len(name.split(".")) > 2 else "?"
        stats = ParseStats(exchange, data_type)
        try:
            records = parse_file(fpath)
            stats.total = len(records)
            for rec in records:
                stats.add_record(rec)
            all_records.extend(records)
        except Exception as e:
            stats.add_error(str(e))
        all_stats.append(stats)
    return all_records, [s.stats_summary for s in all_stats]


def merge_by_code_date(records: list[dict], date_str: str) -> list[dict]:
    """Merge records with same (code, date), filling None from either."""
    merged: dict[str, dict] = {}
    for rec in records:
        key = f"{rec.get('code')}|{rec.get('date')}"
        if key in merged:
            existing = merged[key]
            for k, v in rec.items():
                if v is not None and existing.get(k) is None:
                    if k not in ("code", "date"):
                        existing[k] = v
        else:
            merged[key] = dict(rec)
    # DCE 品种级 maxHand → 合约级 maxoq 传播
    # 品种级 TradingParam.json 按 varietyId 返回 maxHand(如 l-F: maxHand=1000)
    # 解析后 code 为 l.DCE / l-f.DCE(无合约月份数字),合约级如 l2607.DCE / l2607f.DCE
    # 按字母前缀(品种代码 + 可选后缀)匹配传播
    dce_variety: dict[str, dict] = {}
    for rec in merged.values():
        code = rec.get("code", "")
        suffix = code.split(".")[-1] if "." in code else ""
        if suffix.upper() != "DCE":
            continue
        raw = code.split(".")[0]
        if not any(c.isdigit() for c in raw):
            prefix = "".join(c for c in raw if c.isalpha()).upper()
            if prefix:
                dce_variety[prefix] = dict(rec)
    if dce_variety:
        for rec in merged.values():
            code = rec.get("code", "")
            suffix = code.split(".")[-1] if "." in code else ""
            if suffix.upper() != "DCE":
                continue
            raw = code.split(".")[0]
            if any(c.isdigit() for c in raw):
                prefix = "".join(c for c in raw if c.isalpha()).upper()
                if prefix in dce_variety:
                    vrec = dce_variety[prefix]
                    for k, v in vrec.items():
                        if (
                            v is not None
                            and rec.get(k) is None
                            and k not in ("code", "date")
                        ):
                            rec[k] = v

    # OHLC 0 → None(通过 Modifier 做数据清洗)
    for rec in merged.values():
        zero_price_to_none(rec)

    # Compute maxup/maxdown when both pre_settle and _limit_pct are available
    tick_map: dict[str, float] = {}
    for rec in merged.values():
        code = rec.get("code", "")
        pre_settle = rec.get("pre_settle") or rec.get("settle")
        limit_pct = rec.get("_limit_pct")
        limit_pct_down = rec.get("_limit_pct_down", limit_pct)
        if pre_settle is not None and limit_pct is not None:
            if rec.get("maxup") is None:
                if code.endswith((".SHF", ".INE")):
                    # SHFE/INE: 向下取整 via Modifier
                    if not tick_map:
                        tick_map = _load_tick_map(date_str)
                    if tick_map:
                        product = "".join(
                            c.lower() for c in code.split(".")[0]
                            if c.isalpha()
                        )
                        tick = rec.get("_tick_size") or tick_map.get(product)
                        if tick and tick > 0:
                            rec["maxup"], rec["maxdown"] = (
                                calc_shfe_ine_limit_prices(
                                    pre_settle, limit_pct,
                                    limit_pct_down, tick, code
                                )
                            )
                else:
                    # CZCE and others: CZCE-style rounding
                    tick = rec.get("_tick_size")
                    raw_up = pre_settle * (1 + limit_pct)
                    raw_down = pre_settle * (1 - limit_pct_down)
                    rec["maxup"] = (
                        safe_ceil(raw_up, tick)
                        if tick else round(raw_up, 2)
                    )
                    rec["maxdown"] = (
                        safe_floor(raw_down, tick)
                        if tick else round(raw_down, 2)
                    )
        # Clean up internal fields
        for key in ("_limit_pct", "_limit_pct_down", "pre_settle",
                    "_tick_size"):
            rec.pop(key, None)

    return list(merged.values())

# ══════════════════════════════════════════════════════
#  LLM 提取 maxoq / minoq
# ══════════════════════════════════════════════════════

from mxz_utils.logging_config import get_logger

_llm_logger = get_logger(
    name="AnalyseAnnouncementsService",
    level="DEBUG",
    dirpath_logs=str(Path(os.environ["LOG_DIR"])),
    logfile_basename="AnalyseAnnouncementsService",
)

_DS_BASE = "https://api.deepseek.com"
_DS_KEY = os.environ["DEEPSEEK_API_KEY"]
_DS_MODEL = "deepseek-v4-flash"

# 附件下载失败累积列表
_att_failures: list[dict] = []


def clear_attachment_failures():
    """清空附件失败列表(每次运行前调用)。"""
    _att_failures.clear()


def get_attachment_failures() -> list[dict]:
    """获取附件下载失败列表。"""
    return list(_att_failures)


_ZHIPU_KEY = os.environ["ZHIPU_API_KEY"]
_ZHIPU_URL = os.environ["ZHIPU_BASE_URL"]
_ZHIPU_VISION_MODEL = os.environ["ZHIPU_VISION_MODEL"]

_PROMPT_DIR = Path(__file__).parent / "prompts"


def _load_prompt(name: str) -> str:
    """加载 prompts/ 目录下的提示词模板。"""
    return (_PROMPT_DIR / name).read_text(encoding="utf-8")


def _extract_links(html, page_url):
    from urllib.parse import urljoin, unquote
    soup = BeautifulSoup(html, "html.parser")
    results = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        ext = Path(href).suffix.lower()
        if ext in (".pdf", ".xls", ".xlsx"):
            abs_url = urljoin(page_url, href)
            # 从 URL 提取纯数字/字母 ID(去除中文和路径分隔符)
            raw_name = unquote(abs_url.split("/")[-1].split("?")[0])
            url_id = re.sub(r"[^a-zA-Z0-9_.-]", "", Path(raw_name).stem)
            if not url_id:
                url_id = str(abs(hash(abs_url)) % 10**8)
            results.append({"url": abs_url, "url_id": url_id, "ext": ext})
    return results


async def _download(
    url, save_dir, exchange="", date_str="",
    url_id="", session=None,
):
    """下载附件,文件名格式: {exchange}_{date}_{url_id}.{ext}"""
    ext = Path(url).suffix or ".pdf"
    if exchange and date_str and url_id:
        fname = f"{exchange}_{date_str}_{url_id}{ext}"
    else:
        fname = Path(url).name.split("?")[0]
        if not fname or "." not in fname:
            fname = f"att_{abs(hash(url)) % 10**8}{ext}"
    local = save_dir / fname
    if local.exists():
        return local
    save_dir.mkdir(parents=True, exist_ok=True)
    async with session.get(
        url, timeout=aiohttp.ClientTimeout(total=30),
        headers={"User-Agent": "Mozilla/5.0"},
    ) as r:
        r.raise_for_status()
        tmp = local.with_suffix(local.suffix + ".tmp")
        tmp.write_bytes(await r.read())
        tmp.rename(local)
    return local


def _pdf_to_text(pdf_path):
    import fitz, base64, requests as _req
    old_warnings = fitz.TOOLS.mupdf_display_warnings(False)
    try:
        with fitz.open(str(pdf_path)) as doc:
            parts = []
            for p in range(min(len(doc), 10)):
                try:
                    pix = doc[p].get_pixmap(dpi=150)
                    b64 = base64.b64encode(pix.tobytes("png")).decode()
                    r = _req.post(_ZHIPU_URL,
                        headers={"Authorization": f"Bearer {_ZHIPU_KEY}",
                                 "Content-Type": "application/json"},
                        json={
                        "model": _ZHIPU_VISION_MODEL,
                        "messages": [{"role": "user",
                            "content": [{"type": "image_url",
                             "image_url": {"url": f"data:image/png;base64,{b64}"}},
                            {"type": "text", "text": (
        "提取此页中与下单量、开仓手数、交易限额相关的文字,逐字输出。"
    )}
                        ]}], "max_tokens": 1024, "temperature": 0}, timeout=60)
                    parts.append(r.json()["choices"][0]["message"]["content"]
                                if r.status_code == 200 else "")
                except Exception:
                    pass
            page_count = min(len(doc), 10)
    finally:
        fitz.TOOLS.mupdf_display_warnings(old_warnings)
    result = "\n".join(parts)
    _llm_logger.debug(
    "PDF_PARSED %s: %s pages, %s chars",
    pdf_path.name, page_count, len(result),
)
    return result


def _xlsx_to_text(path):
    import openpyxl
    try:
        with openpyxl.load_workbook(str(path), read_only=True, data_only=True) as wb:
            out = []
            for sn in wb.sheetnames[:2]:
                ws = wb[sn]
                rows = [",".join(str(v) for v in r if v is not None)
                        for r in list(ws.iter_rows(values_only=True))[:30]
                        if any(r)
                    ]
                if rows:
                    out.append(f"[Sheet:{sn}]\n" + "\n".join(rows))
            return "\n".join(out)
    except Exception:
        return ""


def _process_html(html, page_url=""):
    """同步 HTML 预处理:提取正文 + 附件链接。放 executor 执行。"""
    soup = BeautifulSoup(html, "html.parser")
    for t in soup(["script", "style", "nav", "header", "footer",
                   "noscript", "iframe"]):
        t.decompose()
    body = None
    for cls in ["TRS_Editor", "article-content", "articleContent"]:
        el = soup.find(class_=re.compile(cls, re.I))
        if el:
            body = el
            break
    text = str(body) if body else soup.get_text()
    for mark in ["第一条", "第一章", "第1条", "第1章",
                 "各会员单位", "各会员"]:
        idx = text.find(mark)
        if idx > 500:
            text = text[idx:]
            break
    links = _extract_links(html, page_url) if page_url else []
    return text, links


def _pad_eff_date(eff_date: str) -> str:
    """补位 effective_date: YYYY → YYYYMMDD, YYYYMM → YYYYMMDD.
    如 eff_date 为空则返回 'YYYYMMDD'."""
    d = eff_date.strip().replace("-", "")
    if not d:
        return "YYYYMMDD"
    if len(d) >= 8:
        return d[:8]
    if len(d) == 6:
        return f"{d}DD"
    if len(d) == 4:
        return f"{d}MMDD"
    return "YYYYMMDD"


async def parse_announcement_fields(
    text: str, links: list, exchange: str, title: str,
    publish_date: str, category: str = "",
    attachment_dir=None, session=None,
) -> list[dict]:
    """异步 LLM 调用解析 minoq/maxoq。text 为预处理后的正文。"""
    # 附件下载
    atext = ""
    if links and attachment_dir:
        for link in links:
            try:
                local = await _download(link["url"], attachment_dir,
                                 exchange, publish_date, link["url_id"],
                                 session=session)
            except BaseException as e:
                _att_failures.append({
                    "exchange": exchange,
                    "title": title,
                    "url": link["url"],
                    "error": str(e),
                })
                _llm_logger.warning(
                    "附件下载失败 [%s] %s → %s: %s",
                    exchange, title[:60], link["url"], e, exc_info=True
                )
                continue
            if local:
                try:
                    t = (_pdf_to_text(local) if link["ext"] == ".pdf"
                         else _xlsx_to_text(local))
                    atext += f"\n[附件:{local.name}]\n{t}\n"
                except Exception:
                    pass

    full_text = text + atext

    prompt = (
        _load_prompt("extract_fields.txt")
        + f"\n交易所:{exchange} 发布日期:{publish_date} 类别:{category}"
        + f"\n\n公告内容:\n\n{full_text}"
    )

    last_error = None
    for retry_i in range(5):
        try:
            if retry_i > 0:
                delay = 2 ** retry_i
                _llm_logger.debug(
            "DS_RETRY [%s] %s: 第 %s/5 次, 等待 %ss",
            exchange, title[:80], retry_i + 1, delay,
        )
                await asyncio.sleep(delay)
            _llm_logger.debug("DS_REQ [%s] %s", exchange, title[:80])
            async with session.post(
                f"{_DS_BASE}/chat/completions",
                headers={"Authorization": f"Bearer {_DS_KEY}",
                         "Content-Type": "application/json"},
                json={"model": _DS_MODEL,
                      "messages": [{"role": "user", "content": prompt}],
                      "response_format": {"type": "json_object"},
                      "temperature": 0, "max_tokens": 393216},
                timeout=aiohttp.ClientTimeout(
                total=600, connect=30, sock_read=300
            ),
            ) as r:
                r.raise_for_status()
                raw = (await r.json())["choices"][0]["message"]["content"]
            _llm_logger.info("DS_RAW [%s] %s: %s", exchange, title[:80], raw)
            break
        except (aiohttp.ClientConnectionError, asyncio.TimeoutError) as e:
            last_error = e
            _llm_logger.warning(
                "DS_ERR [%s] %s: %s",
                exchange, title[:80], e
            )
        except BaseException as e:
            _llm_logger.warning(
                "DS_ERR [%s] %s: %s",
                exchange, title[:80], e, exc_info=True
            )
            raise
    else:
        _llm_logger.warning(
            "DS_ERR [%s] %s: %s",
            exchange, title[:80], last_error, exc_info=True
        )
        raise last_error

    raw = raw.strip()
    data = json.loads(raw)
    items = data if isinstance(data, list) else data.get("items", [])
    needs_review = (
                    False if isinstance(data, list)
                    else bool(data.get("needs_review", False))
                )
    results = []
    for item in items:
        # 除 effective_date 外全部必填
        try:
            pc = item["product_code"]
            sid = item["security_id"]
            field = item["field"]
            val = item["value"]
            evidence = item["evidence"]
        except KeyError as e:
            raise KeyError(
                f"LLM 输出缺少字段 {e} [{exchange}] {title[:80]}"
            ) from e
        if field not in ("minoq", "maxoq"):
            continue
        # effective_date: 缺失或不全时用占位字母补位
        eff_date = _pad_eff_date(item.get("effective_date", ""))
        results.append({
            "product_code": str(pc),
            "security_id": str(sid),
            "field": field,
            "value": int(val),
            "effective_date": eff_date,
            "evidence": str(evidence),
        })
    return results, needs_review
